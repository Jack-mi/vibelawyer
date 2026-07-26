"""阅卷目录 Excel 生成器.

三个工作表:
  1. 阅卷目录 —— 卷宗名称/页码/所含文件/文书类型/笔录时间/备注
  2. 案件信息 —— 结构化基本信息表
  3. 证据索引 —— 全部记录与其来源引用的追溯总表（落实可回溯约束）
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..workspace import CaseWorkspace


_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")
_TITLE_FONT = Font(bold=True, size=14)


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_body(ws, start_row: int, ncols: int, end_row: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER
            cell.alignment = _WRAP


def _set_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_catalog_xlsx(ws: CaseWorkspace, out_dir: Path) -> Path:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ---- Sheet 1: 阅卷目录 ----
    s1 = wb.active
    s1.title = "阅卷目录"
    s1["A1"] = f"{ws.case_name or '案件'} 阅卷目录"
    s1["A1"].font = _TITLE_FONT
    s1.merge_cells("A1:F1")
    headers = ["卷宗名称", "页码", "本卷所含文件名称", "文书类型", "笔录时间", "备注"]
    s1.append([])
    s1.append(headers)
    _style_header(s1, 3, len(headers))
    row = 4
    if ws.catalog:
        for e in ws.catalog:
            s1.append([e.volume_name, e.page_range, e.file_name, e.doc_type, e.record_time, e.note])
            row += 1
    else:
        s1.append(["（暂无）"] * 6)
        row += 1
    _style_body(s1, 4, len(headers), row - 1)
    _set_widths(s1, [16, 12, 32, 16, 16, 20])
    s1.freeze_panes = "A4"

    # ---- Sheet 2: 案件信息 ----
    s2 = wb.create_sheet("案件信息")
    s2["A1"] = "案件基本信息表"
    s2["A1"].font = _TITLE_FONT
    s2.merge_cells("A1:B1")
    s2.append([])
    s2.append(["项目", "内容"])
    _style_header(s2, 3, 2)
    ind = ws.indictment
    rows = [
        ["1. 代理主体（被告人）", ws.defendant or "—"],
        ["2. 涉嫌罪名", ws.charge or "—"],
        ["3. 起诉书/起诉意见书指控事实", ind.full_text_summary or "—"],
        ["4. 涉案金额", ws.total_amount or "—"],
        ["5. 卷宗数量", f"{ws.volume_count} 册"],
        ["  卷宗清单", "；".join(f"《{v}》" for v in ws.known_volumes()) or "—"],
        ["6. 指控事实分笔数", str(len(ind.facts))],
        ["7. 被告人供述笔录数", str(len(ws.defendant_statements))],
        ["8. 同案人供述笔录数", str(len(ws.codefendant_statements))],
        ["9. 证人证言笔录数", str(len(ws.witness_statements))],
        ["10. 程序性文书数", str(len(ws.procedural_docs))],
        ["11. 书证数", str(len(ws.documentary_evidence))],
        ["起诉书制作机关", ind.issuer or "—"],
        ["起诉书落款日期", ind.issue_date or "—"],
        ["适用法律条文", "；".join(ind.legal_basis) or "—"],
    ]
    for r in rows:
        s2.append(r)
    _style_body(s2, 4, 2, 3 + len(rows))
    _set_widths(s2, [22, 70])
    s2.freeze_panes = "A4"

    # ---- Sheet 3: 证据索引（追溯总表）----
    s3 = wb.create_sheet("证据索引")
    s3["A1"] = "证据与事实来源追溯总表（每一处结论可回溯到具体卷宗页码）"
    s3["A1"].font = _TITLE_FONT
    s3.merge_cells("A1:F1")
    s3.append([])
    h3 = ["部分", "记录", "卷宗", "页码", "关键内容", "来源引用"]
    s3.append(h3)
    _style_header(s3, 3, len(h3))
    idx = 4

    def add(part: str, label: str, vol: str, ps: int, pe: int, content: str, cite_text: str) -> None:
        nonlocal idx
        s3.append([part, label, vol or "—", f"{ps}-{pe}" if ps else "—", content, cite_text])
        idx += 1

    # 当事人
    if ws.party.name:
        add("一.当事人", ws.party.name, ws.party.source.volume if ws.party.source else "",
            ws.party.source.page_start if ws.party.source else 0,
            ws.party.source.page_end if ws.party.source else 0,
            f"{ws.party.name}，{ws.party.position or ws.party.occupation}",
            (ws.party.source.render() if ws.party.source else "—"))
    # 起诉书
    if ws.indictment.doc_type:
        add("二.起诉书", ws.indictment.doc_type, ws.indictment.source.volume if ws.indictment.source else "",
            ws.indictment.source.page_start if ws.indictment.source else 0,
            ws.indictment.source.page_end if ws.indictment.source else 0,
            f"{ws.indictment.charge}，金额{ws.indictment.total_amount}",
            (ws.indictment.source.render() if ws.indictment.source else "—"))
    for f in ws.indictment.facts:
        add("二.指控事实", f"第{f.index}笔", f.source.volume if f.source else "",
            f.source.page_start if f.source else 0, f.source.page_end if f.source else 0,
            f.description, (f.source.render() if f.source else "—"))
    for s in ws.defendant_statements:
        add("三.被告人供述", s.person, s.volume, s.page_start, s.page_end, s.content_summary[:60],
            (s.source.render() if s.source else "—"))
    for s in ws.codefendant_statements:
        add("四.同案人供述", s.person, s.volume, s.page_start, s.page_end, s.content_summary[:60],
            (s.source.render() if s.source else "—"))
    for s in ws.witness_statements:
        add("五.证人证言", s.person, s.volume, s.page_start, s.page_end, s.content_summary[:60],
            (s.source.render() if s.source else "—"))
    for d in ws.procedural_docs:
        add("六.程序性文书", d.doc_type, d.volume, d.page_start, d.page_end, d.content_summary[:60],
            (d.source.render() if d.source else "—"))
    for e in ws.documentary_evidence:
        add("七.书证", e.name, e.volume, e.page_start, e.page_end, e.content_summary[:60],
            (e.source_ref.render() if e.source_ref else "—"))
    _style_body(s3, 4, len(h3), idx - 1)
    _set_widths(s3, [14, 16, 16, 10, 50, 22])
    s3.freeze_panes = "A4"

    safe = (ws.case_name or "case").replace("/", "_").replace(" ", "")
    path = out_dir / f"{safe}_阅卷目录.xlsx"
    wb.save(str(path))
    return path
